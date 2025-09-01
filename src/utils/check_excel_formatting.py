import pandas as pd
from openpyxl import load_workbook

# Check the new file with strikethrough
try:
    wb = load_workbook(
        "d:/Coding/fantasy drafting/output/redraft_12team_cheatsheet.xlsx"
    )
    ws = wb["Draft Board"]

    print("✅ File loaded successfully")
    print(f"Conditional formatting rules: {len(ws.conditional_formatting)}")

    for i, rule in enumerate(ws.conditional_formatting):
        print(
            f'Rule {i+1}: {rule.cells} - Formula: {rule.rules[0].formula if rule.rules else "No formula"}'
        )

    # Check column headers
    headers = []
    for col in range(1, 12):
        cell_value = ws.cell(row=1, column=col).value
        headers.append(cell_value)

    print(f"\nColumn headers: {headers}")

    # Find the Drafted column
    drafted_col_name = "Drafted (Enter 'X' to cross out row)"
    if drafted_col_name in headers:
        col_index = headers.index(drafted_col_name) + 1
        print(f"Drafted column is: Column {col_index} ({chr(64 + col_index)})")
    else:
        print("❌ Drafted column not found!")

    # Check a few sample rows
    print("\nSample data (first 3 rows):")
    for row in range(2, 5):
        row_data = []
        for col in range(1, 10):
            cell_value = ws.cell(row=row, column=col).value
            row_data.append(str(cell_value)[:20] if cell_value else "")
        print(f"Row {row}: {row_data}")

except Exception as e:
    print(f"❌ Error: {e}")
    import traceback

    traceback.print_exc()
