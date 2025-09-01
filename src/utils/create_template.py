from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Create a simple template file for conditional formatting setup
wb = load_workbook("../output/dynasty_superflex_cheatsheet.xlsx")
ws = wb["Draft Board"]

print("Creating conditional formatting template...")

# Clear most data and keep only headers and a few sample rows
# This makes it easier to set up conditional formatting

# Add clear setup instructions at the top
ws.insert_rows(1)
ws.cell(row=1, column=1).value = "CONDITIONAL FORMATTING SETUP TEMPLATE"
ws.cell(row=1, column=1).font = Font(bold=True, size=14)
ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

# Merge cells for the title
from openpyxl.utils import get_column_letter

ws.merge_cells(f"A1:{get_column_letter(ws.max_column)}1")

# Add setup instructions
ws.cell(row=2, column=1).value = "INSTRUCTIONS:"
ws.cell(row=2, column=1).font = Font(bold=True)

instructions = [
    "1. Select columns A:H (click column A header, hold Shift, click column H header)",
    "2. Go to Home > Conditional Formatting > New Rule",
    '3. Select "Use a formula to determine which cells to format"',
    '4. Enter formula: =$I2="X"',
    '5. Click Format > Font tab > Check "Strikethrough"',
    "6. Click OK twice",
    '7. Test by typing "X" in column I of any row',
    "",
    "ORIGINAL DATA STARTS BELOW:",
]

for i, instruction in enumerate(instructions, 3):
    ws.cell(row=i, column=1).value = instruction
    if i <= 9:  # Make instructions bold
        ws.cell(row=i, column=1).font = Font(bold=True)

# Save the template
wb.save("../output/dynasty_superflex_cheatsheet_TEMPLATE.xlsx")
print("Template file created: dynasty_superflex_cheatsheet_TEMPLATE.xlsx")
print("This file has clear instructions for setting up conditional formatting!")
