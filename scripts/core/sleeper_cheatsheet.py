#!/usr/bin/env python3
"""
Sleeper-Tailored Fantasy Cheat Sheet Builder
============================================

What this does
--------------
- Loads **Sleeper ADP** (CSV export is easiest) and your **projections** (CSV).
- Applies your **custom scoring** (incl. bonuses) for offense and optional IDP.
- Computes **VORP** with **Superflex-aware** QB replacement.
- Creates **tiers** using configurable point-drop thresholds.
- Exports a **color-coded Excel** file with Overall + per-position sheets.

How to run (from a terminal / VS Code)
--------------------------------------
1) Install deps:  pip install -r requirements.txtk
2) Put your CSVs in the same folder as this script (defaults below), or update paths in CONFIG.
   - sleeper_adp.csv              -> Sleeper ADP export (see accepted column names below)
   - offense_projections.csv      -> Projections for QB/RB/WR/TE
   - idp_projections.csv          -> (Optional) Projections for DL/LB/DB
3) Update the CONFIG_FILE and CONFIG_PROFILE variables at the top of this script
4) Run:  python sleeper_cheatsheet.py
5) Open the generated: fantasy_cheatsheet.xlsx

Accepted ADP CSV columns (any that match will be used):
- Player name:  'player', 'player_name', 'name', 'display_name'
- Position:     'position', 'pos'
- Team:         'team', 'team_abbr', 'team_code', 'nfl_team'
- ADP:          'adp', 'average_draft_position', 'avg_pick'

Notes on Sleeper API (optional)
-------------------------------
This script *can* fetch live ADP if internet access is available.
If you prefer, set CONFIG['sleeper_api']['use_api'] = True and adjust params.
Endpoints and formats can change—CSV import is the most reliable approach.

"""

from __future__ import annotations

import json
import os
from typing import Dict, List

import numpy as np
import pandas as pd

# Excel formatting
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule

# =============================================================================
# CONFIGURATION - Set your desired config file and profile here
# =============================================================================
CONFIG_FILE = os.path.join(
    os.path.dirname(__file__), "../../configs/config_redraft.json"
)  # Path to your config file
CONFIG_PROFILE = "redraft_12team"  # Profile name within the config file

# Quick config presets (uncomment the one you want to use):
# CONFIG_FILE, CONFIG_PROFILE = "list", "all"                     # List all configs
# CONFIG_FILE, CONFIG_PROFILE = "../configs/config.json", "default"           # Standard league
# CONFIG_FILE, CONFIG_PROFILE = "../configs/config_redraft.json", "redraft_12team"   # Redraft league
# CONFIG_FILE, CONFIG_PROFILE = "../configs/config_dynasty.json", "superflex_dynasty"   # Dynasty league
# CONFIG_FILE, CONFIG_PROFILE = "../configs/config_superflex.json", "superflex_12team"  # Superflex league
# =============================================================================

try:
    import requests
except Exception:
    requests = None  # in case user doesn't need API fetch


# -------------------------------
# 1) CONFIG LOADING
# -------------------------------


def load_config(config_file: str = "config.json", config_name: str = "default") -> dict:
    """
    Load configuration from JSON file.

    Args:
        config_file: Path to JSON config file
        config_name: Name of the configuration to load (e.g., 'default', 'redraft_12team')

    Returns:
        Configuration dictionary
    """
    if not os.path.exists(config_file):
        raise FileNotFoundError(f"Configuration file not found: {config_file}")

    with open(config_file, "r") as f:
        configs = json.load(f)

    if config_name not in configs:
        available = list(configs.keys())
        raise ValueError(
            f"Configuration '{config_name}' not found. Available: {available}"
        )

    return configs[config_name]


def get_available_configs(config_file: str = "config.json") -> list:
    """Get list of available configuration names."""
    if not os.path.exists(config_file):
        return []

    with open(config_file, "r") as f:
        configs = json.load(f)

    return list(configs.keys())


def list_all_available_configs():
    """Helper function to list all configurations from all config files."""
    config_files = [
        "config.json",
        "config_redraft.json",
        "config_dynasty.json",
        "config_superflex.json",
        "config_example.json",
    ]

    print("📋 Available Configurations:")
    print("=" * 50)

    for config_file in config_files:
        if os.path.exists(config_file):
            configs = get_available_configs(config_file)
            if configs:
                print(f"\n{config_file}:")
                for config_name in configs:
                    print(f"  - {config_name}")
                    print(
                        f'    Usage: CONFIG_FILE, CONFIG_PROFILE = "{config_file}", "{config_name}"'
                    )
    print("\n" + "=" * 50)
    print(
        "💡 Copy one of the usage lines above to the CONFIG section at the top of this script."
    )


# -------------------------------
# 2) COLUMN MAPPINGS
# -------------------------------

# Flexible projections columns (rename if needed)
OFFENSE_MAP = {
    "player": [
        "player",
        "player_name",
        "name",
        "display_name",
        "PLAYER NAME",
        "Player",
    ],
    "position": ["position", "pos", "POS", "Pos"],
    "team": ["team", "team_abbr", "team_code", "nfl_team", "TEAM", "Team"],
    "age": ["age"],
    "rank": ["RK", "rank", "overall_rank"],
    "tier": ["TIERS", "tier"],
    # passing
    "pass_yards": ["pass_yards", "pass_yd", "pass_yds", "py", "YDS"],
    "pass_td": ["pass_td", "pass_tds", "TDS"],
    "pass_int": ["pass_int", "ints", "INTS"],
    "pass_2pt": ["pass_2pt", "pass_two_pt"],
    "pass_comp_25p_events": ["pass_comp_25p_events"],  # optional expected counts
    # rushing
    "rush_yards": ["rush_yards", "rush_yd", "rush_yds", "ry", "YDS.1"],
    "rush_td": ["rush_td", "rush_tds", "TDS.1"],
    "rush_2pt": ["rush_2pt", "rush_two_pt"],
    "rush_att_20p_events": ["rush_att_20p_events"],  # optional expected counts
    "rush_first_down": ["rush_first_down"],
    # receiving
    "receptions": ["receptions", "rec", "recpt", "REC"],
    "rec_yards": ["rec_yards", "rec_yd", "rec_yds", "ryds"],
    "rec_td": ["rec_td", "rec_tds"],
    "rec_2pt": ["rec_2pt", "rec_two_pt"],
    "rec_first_down": ["rec_first_down"],
    # long reception buckets (optional)
    "rec_10_19": ["rec_10_19"],
    "rec_20_29": ["rec_20_29"],
    "rec_30_39": ["rec_30_39"],
    "rec_40_plus": ["rec_40_plus"],
    "rec_40_plus_td": ["rec_40_plus_td"],
    "rec_50_plus_td": ["rec_50_plus_td"],
    # per-game bonus events (optional)
    "pg_100_199_rush": ["pg_100_199_rush"],
    "pg_200_plus_rush": ["pg_200_plus_rush"],
    "pg_100_199_rec": ["pg_100_199_rec"],
    "pg_200_plus_rec": ["pg_200_plus_rec"],
    "pg_100_199_combo": ["pg_100_199_combo"],
    "pg_200_plus_combo": ["pg_200_plus_combo"],
    "pg_300_399_pass": ["pg_300_399_pass"],
    "pg_400_plus_pass": ["pg_400_plus_pass"],
}

IDP_MAP = {
    "player": ["player", "player_name", "name", "display_name", "PLAYER NAME", "Name"],
    "position": ["position", "pos", "POS", "Pos"],  # DL/LB/DB
    "team": ["team", "team_abbr", "team_code", "nfl_team", "TEAM", "Team"],
    "age": ["age"],
    "rank": ["RK", "rank", "overall_rank", "#"],
    "tier": ["TIERS", "tier"],
    "tackle_solo": ["tackle_solo", "solo", "Tackles Solo"],
    "tackle_ast": ["tackle_ast", "ast", "Tackles Ast"],
    "tackles_total": ["tackles_total", "tackles", "Tackles"],
    "sack": ["sack", "sacks", "Sacks"],
    "int": ["int", "ints", "Ints"],
    "ff": ["ff", "forced_fumbles", "Fum Forc"],
    "fr": ["fr", "fumble_rec", "Fum Rec"],
    "def_td": ["def_td", "def_tds", "TD Ret"],
    "pd": ["pd", "passes_defended", "pass_def", "Pass Def"],
    "safety": ["safety", "safeties", "Saf"],
    "blk": ["blk", "blocked_kick"],
}

ADP_MAP = {
    "player": ["player", "player_name", "name", "display_name", "Player"],
    "position": ["position", "pos"],
    "team": ["team", "team_abbr", "team_code", "nfl_team"],
    "adp": ["adp", "average_draft_position", "avg_pick", "Overall Rank"],
}


# -------------------------------
# 3) HELPERS AND DATA PROCESSORS
# -------------------------------


class DataProcessor:
    """Handles data processing operations with consistent methods."""

    @staticmethod
    def coalesce_column(df: pd.DataFrame, targets: List[str], default=None):
        """Find first matching column from list of candidates."""
        for t in targets:
            if t in df.columns:
                return df[t]
        return pd.Series([default] * len(df))

    @staticmethod
    def clean_numeric(value):
        """Convert values to numeric, handling commas and invalid data."""
        if pd.isna(value):
            return 0.0
        try:
            return float(str(value).replace(",", ""))
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def to_numeric_columns(
        df: pd.DataFrame, exclude=("player", "position", "team", "age")
    ) -> pd.DataFrame:
        """Convert all columns except specified ones to numeric."""
        df_copy = df.copy()

        for c in df_copy.columns:
            if c not in exclude:
                df_copy[c] = pd.to_numeric(df_copy[c], errors="coerce").fillna(0.0)

        return df_copy


class PositionMapper:
    """Handles position standardization and mapping."""

    POSITION_MAP = {
        "WR": "WR",
        "RB": "RB",
        "QB": "QB",
        "TE": "TE",
        "DE": "DL",
        "DT": "DL",
        "DL": "DL",
        "EDGE": "DL",
        "LB": "LB",
        "ILB": "LB",
        "OLB": "LB",
        "MLB": "LB",
        "CB": "DB",
        "S": "DB",
        "FS": "DB",
        "SS": "DB",
        "DB": "DB",
        "K": "K",
        "DST": "DST",
        "DEF": "DST",
        "D/ST": "DST",
    }

    @classmethod
    def extract_position(cls, pos_str: str) -> str:
        """Extract base position from position strings like 'WR1', 'RB2', 'DE1', etc."""
        if pd.isna(pos_str) or pos_str == "":
            return ""

        import re

        # Remove numbers and common suffixes
        pos = re.sub(r"\d+", "", str(pos_str)).upper()
        return cls.POSITION_MAP.get(pos, pos)


# Legacy functions for backward compatibility
def coalesce_column(df: pd.DataFrame, targets: List[str], default=None):
    return DataProcessor.coalesce_column(df, targets, default)


def extract_position(pos_str: str) -> str:
    return PositionMapper.extract_position(pos_str)


class ScoringEngine:
    """Handles all scoring calculations with clean separation of concerns."""

    @staticmethod
    def assign_points_from_rankings(df: pd.DataFrame, position: str) -> pd.Series:
        """
        When projection stats aren't available, assign points based on rankings.
        Higher-ranked players get more points using a decay function.
        """
        if "rank" in df.columns:
            rank = pd.to_numeric(df["rank"], errors="coerce").fillna(999)

            # Position-specific point ranges
            if position in ["QB"]:
                base_points, decay_factor = 280, 0.05
            elif position in ["RB", "WR"]:
                base_points, decay_factor = 250, 0.04
            elif position in ["TE"]:
                base_points, decay_factor = 200, 0.06
            else:  # IDP positions
                base_points, decay_factor = 180, 0.03

            return base_points * np.exp(-decay_factor * (rank - 1))
        else:
            return pd.Series([100.0] * len(df))

    @staticmethod
    def apply_offense_scoring(
        df: pd.DataFrame, scoring_config: Dict[str, float]
    ) -> pd.Series:
        """Apply offensive scoring to standardized DataFrame."""
        # Ensure required columns exist with default values
        required_cols = [
            "pass_yards",
            "pass_td",
            "pass_int",
            "pass_2pt",
            "rush_yards",
            "rush_td",
            "rush_2pt",
            "receptions",
            "rec_yards",
            "rec_td",
            "rec_2pt",
            "rush_first_down",
            "rec_first_down",
        ]

        # Convert to numeric and fill missing columns
        for col in required_cols:
            if col not in df.columns:
                df[col] = 0.0
            else:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

        # Use .get() with default values to avoid KeyError
        return (
            df.get("pass_yards", pd.Series(0.0, index=df.index))
            * scoring_config.get("pass_yd", 0.0)
            + df.get("pass_td", pd.Series(0.0, index=df.index))
            * scoring_config.get("pass_td", 0.0)
            + df.get("pass_int", pd.Series(0.0, index=df.index))
            * scoring_config.get("pass_int", 0.0)
            + df.get("pass_2pt", pd.Series(0.0, index=df.index))
            * scoring_config.get("pass_2pt", 0.0)
            + df.get("rush_yards", pd.Series(0.0, index=df.index))
            * scoring_config.get("rush_yd", 0.0)
            + df.get("rush_td", pd.Series(0.0, index=df.index))
            * scoring_config.get("rush_td", 0.0)
            + df.get("rush_2pt", pd.Series(0.0, index=df.index))
            * scoring_config.get("rush_2pt", 0.0)
            + df.get("receptions", pd.Series(0.0, index=df.index))
            * scoring_config.get("rec", 0.0)
            + df.get("rec_yards", pd.Series(0.0, index=df.index))
            * scoring_config.get("rec_yd", 0.0)
            + df.get("rec_td", pd.Series(0.0, index=df.index))
            * scoring_config.get("rec_td", 0.0)
            + df.get("rec_2pt", pd.Series(0.0, index=df.index))
            * scoring_config.get("rec_2pt", 0.0)
            + df.get("rush_first_down", pd.Series(0.0, index=df.index))
            * scoring_config.get("first_down_rb", 0.0)
            + df.get("rec_first_down", pd.Series(0.0, index=df.index))
            * scoring_config.get("first_down_wr", 0.0)
            + df.get("pass_comp_25p_events", pd.Series(0.0, index=df.index))
            * scoring_config.get("bonus_25_plus_completions", 0.0)
            + df.get("rush_att_20p_events", pd.Series(0.0, index=df.index))
            * scoring_config.get("bonus_20_plus_carries", 0.0)
            + df.get("rec_10_19", pd.Series(0.0, index=df.index))
            * scoring_config.get("rec_10_19_bonus", 0.0)
            + df.get("rec_20_29", pd.Series(0.0, index=df.index))
            * scoring_config.get("rec_20_29_bonus", 0.0)
            + df.get("rec_30_39", pd.Series(0.0, index=df.index))
            * scoring_config.get("rec_30_39_bonus", 0.0)
            + df.get("rec_40_plus", pd.Series(0.0, index=df.index))
            * scoring_config.get("rec_40_plus_bonus", 0.0)
            + df.get("rec_40_plus_td", pd.Series(0.0, index=df.index))
            * scoring_config.get("rec_40_plus_td_bonus", 0.0)
            + df.get("rec_50_plus_td", pd.Series(0.0, index=df.index))
            * scoring_config.get("rec_50_plus_td_bonus", 0.0)
            + df.get("pg_100_199_rush", pd.Series(0.0, index=df.index))
            * scoring_config.get("bonus_100_199_rush_game", 0.0)
            + df.get("pg_200_plus_rush", pd.Series(0.0, index=df.index))
            * scoring_config.get("bonus_200_plus_rush_game", 0.0)
            + df.get("pg_100_199_rec", pd.Series(0.0, index=df.index))
            * scoring_config.get("bonus_100_199_rec_game", 0.0)
            + df.get("pg_200_plus_rec", pd.Series(0.0, index=df.index))
            * scoring_config.get("bonus_200_plus_rec_game", 0.0)
            + df.get("pg_100_199_combo", pd.Series(0.0, index=df.index))
            * scoring_config.get("bonus_100_199_combo_game", 0.0)
            + df.get("pg_200_plus_combo", pd.Series(0.0, index=df.index))
            * scoring_config.get("bonus_200_plus_combo_game", 0.0)
            + df.get("pg_300_399_pass", pd.Series(0.0, index=df.index))
            * scoring_config.get("bonus_300_399_pass_game", 0.0)
            + df.get("pg_400_plus_pass", pd.Series(0.0, index=df.index))
            * scoring_config.get("bonus_400_plus_pass_game", 0.0)
        )

    @staticmethod
    def apply_idp_scoring(
        df: pd.DataFrame, scoring_config: Dict[str, float]
    ) -> pd.Series:
        """Apply IDP scoring to standardized DataFrame."""
        df = DataProcessor.to_numeric_columns(df)

        # Handle different tackle scoring systems
        if "tackle_solo" in df.columns and "tackle_ast" in df.columns:
            tackle_points = df.get(
                "tackle_solo", pd.Series(0.0, index=df.index)
            ) * scoring_config.get("tackle_solo", 0.0) + df.get(
                "tackle_ast", pd.Series(0.0, index=df.index)
            ) * scoring_config.get(
                "tackle_ast", 0.0
            )
        elif "tackles_total" in df.columns:
            # Assume roughly 60% solo, 40% assisted for total tackle value
            avg_tackle_value = (
                scoring_config.get("tackle_solo", 0.0) * 0.6
                + scoring_config.get("tackle_ast", 0.0) * 0.4
            )
            tackle_points = (
                df.get("tackles_total", pd.Series(0.0, index=df.index))
                * avg_tackle_value
            )
        else:
            tackle_points = pd.Series(0.0, index=df.index)

        return (
            tackle_points
            + df.get("sack", pd.Series(0.0, index=df.index))
            * scoring_config.get("sack", 0.0)
            + df.get("int", pd.Series(0.0, index=df.index))
            * scoring_config.get("int", 0.0)
            + df.get("ff", pd.Series(0.0, index=df.index))
            * scoring_config.get("ff", 0.0)
            + df.get("fr", pd.Series(0.0, index=df.index))
            * scoring_config.get("fr", 0.0)
            + df.get("def_td", pd.Series(0.0, index=df.index))
            * scoring_config.get("def_td", 0.0)
            + df.get("pd", pd.Series(0.0, index=df.index))
            * scoring_config.get("pd", 0.0)
            + df.get("safety", pd.Series(0.0, index=df.index))
            * scoring_config.get("safety", 0.0)
            + df.get("blk", pd.Series(0.0, index=df.index))
            * scoring_config.get("blk", 0.0)
        )


class DataStandardizer:
    """Handles data standardization with position-aware column mapping."""

    @staticmethod
    def standardize(df: pd.DataFrame, mapping: Dict[str, List[str]]) -> pd.DataFrame:
        """Basic column standardization using mapping dictionary."""
        out = pd.DataFrame()
        for k, candidates in mapping.items():
            out[k] = DataProcessor.coalesce_column(df, candidates, 0)
        return out

    @staticmethod
    def standardize_offense_position_aware(df: pd.DataFrame) -> pd.DataFrame:
        """
        Position-aware column mapping for offense projections.
        Handles cases where YDS/TDS could be passing or receiving stats.
        """
        out = pd.DataFrame()

        # Basic columns that don't need position logic
        basic_mapping = {
            "player": [
                "player",
                "player_name",
                "name",
                "display_name",
                "PLAYER NAME",
                "Player",
            ],
            "position": ["position", "pos", "POS", "Pos"],
            "team": ["team", "team_abbr", "team_code", "nfl_team", "TEAM", "Team"],
            "age": ["age"],
            "rank": ["RK", "rank", "overall_rank"],
            "tier": ["TIERS", "tier"],
        }

        for k, candidates in basic_mapping.items():
            out[k] = DataProcessor.coalesce_column(df, candidates, 0)

        # Get positions to determine stat mapping
        positions = DataProcessor.coalesce_column(df, basic_mapping["position"])

        # Initialize all stat columns to 0
        stat_cols = [
            "pass_yards",
            "pass_td",
            "pass_int",
            "pass_2pt",
            "pass_comp_25p_events",
            "rush_yards",
            "rush_td",
            "rush_2pt",
            "rush_att_20p_events",
            "rush_first_down",
            "receptions",
            "rec_yards",
            "rec_td",
            "rec_2pt",
            "rec_first_down",
            "rec_10_19",
            "rec_20_29",
            "rec_30_39",
            "rec_40_plus",
            "rec_40_plus_td",
            "rec_50_plus_td",
            "pg_100_199_rush",
            "pg_200_plus_rush",
            "pg_100_199_rec",
            "pg_200_plus_rec",
            "pg_100_199_combo",
            "pg_200_plus_combo",
            "pg_300_399_pass",
            "pg_400_plus_pass",
        ]

        for col in stat_cols:
            out[col] = pd.Series([0.0] * len(df))

        # Position-specific stat mapping
        for i, pos in enumerate(positions):
            pos_clean = PositionMapper.extract_position(str(pos))
            DataStandardizer._map_position_stats(df, out, i, pos_clean)

        return out

    @staticmethod
    def _map_position_stats(
        df: pd.DataFrame, out: pd.DataFrame, index: int, position: str
    ):
        """Map stats based on position for a specific row."""
        clean_numeric = DataProcessor.clean_numeric

        if position == "QB":
            # For QBs: YDS = passing yards, TDS = passing TDs
            if "YDS" in df.columns:
                out.loc[index, "pass_yards"] = clean_numeric(df.iloc[index]["YDS"])
            if "TDS" in df.columns:
                out.loc[index, "pass_td"] = clean_numeric(df.iloc[index]["TDS"])
            if "INTS" in df.columns:
                out.loc[index, "pass_int"] = clean_numeric(df.iloc[index]["INTS"])
            if "YDS.1" in df.columns:
                out.loc[index, "rush_yards"] = clean_numeric(df.iloc[index]["YDS.1"])
            if "TDS.1" in df.columns:
                out.loc[index, "rush_td"] = clean_numeric(df.iloc[index]["TDS.1"])

        elif position in ["WR", "TE"]:
            # For WR/TE: YDS = receiving yards, TDS = receiving TDs
            if "REC" in df.columns:
                out.loc[index, "receptions"] = clean_numeric(df.iloc[index]["REC"])
            if "YDS" in df.columns:
                out.loc[index, "rec_yards"] = clean_numeric(df.iloc[index]["YDS"])
            if "TDS" in df.columns:
                out.loc[index, "rec_td"] = clean_numeric(df.iloc[index]["TDS"])
            if "YDS.1" in df.columns:
                out.loc[index, "rush_yards"] = clean_numeric(df.iloc[index]["YDS.1"])
            if "TDS.1" in df.columns:
                out.loc[index, "rush_td"] = clean_numeric(df.iloc[index]["TDS.1"])

        elif position == "RB":
            # For RB: YDS.1 = rushing yards, TDS.1 = rushing TDs, YDS = receiving yards, TDS = receiving TDs
            if "REC" in df.columns:
                out.loc[index, "receptions"] = clean_numeric(df.iloc[index]["REC"])
            if "YDS" in df.columns:
                out.loc[index, "rec_yards"] = clean_numeric(df.iloc[index]["YDS"])
            if "TDS" in df.columns:
                out.loc[index, "rec_td"] = clean_numeric(df.iloc[index]["TDS"])
            if "YDS.1" in df.columns:
                out.loc[index, "rush_yards"] = clean_numeric(df.iloc[index]["YDS.1"])
            if "TDS.1" in df.columns:
                out.loc[index, "rush_td"] = clean_numeric(df.iloc[index]["TDS.1"])


# Legacy functions for backward compatibility
def standardize(df: pd.DataFrame, mapping: Dict[str, List[str]]) -> pd.DataFrame:
    return DataStandardizer.standardize(df, mapping)


def standardize_offense_position_aware(df: pd.DataFrame) -> pd.DataFrame:
    return DataStandardizer.standardize_offense_position_aware(df)


def to_numeric(
    df: pd.DataFrame, exclude=("player", "position", "team", "age")
) -> pd.DataFrame:
    return DataProcessor.to_numeric_columns(df, exclude)


def apply_offense_scoring(df: pd.DataFrame, sc: Dict[str, float]) -> pd.Series:
    return ScoringEngine.apply_offense_scoring(df, sc)


def apply_idp_scoring(df: pd.DataFrame, sc: Dict[str, float]) -> pd.Series:
    return ScoringEngine.apply_idp_scoring(df, sc)


def assign_points_from_rankings(df: pd.DataFrame, position: str) -> pd.Series:
    return ScoringEngine.assign_points_from_rankings(df, position)


class DraftPriorityCalculator:
    """Handles the complex Draft Priority scoring algorithm."""

    def __init__(self, league_config: dict, draft_strategy_config: dict = None):
        self.teams = league_config["num_teams"]
        self.is_superflex = league_config["superflex"]
        self.starters = league_config["starters"]

        # Use draft strategy config or fall back to statistical thresholds
        self.strategy_config = (
            draft_strategy_config or self._get_default_strategy_config()
        )

    def _get_default_strategy_config(self) -> dict:
        """Default strategy based on statistical thresholds rather than hardcoded names."""
        return {
            "round_bonuses": {
                "rounds_1_2": {
                    "bonus": 50,
                    "criteria": {
                        "WR": {"min_vorp": 50, "min_rank": 5},
                        "QB": {"min_vorp": 80, "min_rank": 3},
                    },
                },
                "round_3": {
                    "bonus": 40,
                    "criteria": {
                        "RB": {"min_vorp": 70, "min_rank": 8},
                        "DL": {"min_vorp": 50, "min_rank": 5},
                    },
                },
                "rounds_4_5": {
                    "bonus": 35,
                    "criteria": {
                        "DL": {"min_vorp": 20, "max_rank": 15},
                        "LB": {"min_vorp": 15, "max_rank": 12},
                        "QB": {"min_vorp": 60, "max_rank": 8},
                        "RB": {"min_vorp": 20, "max_rank": 20},
                        "WR": {"min_vorp": 20, "max_rank": 20},
                    },
                },
                "rounds_6_7": {
                    "bonus": 20,
                    "criteria": {
                        "QB": {"min_vorp": 40, "max_rank": 15},
                        "RB": {"min_vorp": 10, "max_rank": 30},
                        "WR": {"min_vorp": 10, "max_rank": 30},
                    },
                },
                "rounds_8_plus": {
                    "bonus": 12,
                    "criteria": {
                        "DL": {"min_vorp": -10},
                        "LB": {"min_vorp": -10},
                        "DB": {"min_vorp": -10},
                        "QB": {"min_vorp": 20, "max_rank": 25},
                        "RB": {"min_vorp": -10, "max_rank": 50},
                        "WR": {"min_vorp": -10, "max_rank": 50},
                        "TE": {"min_vorp": -10, "max_rank": 30},
                    },
                },
            }
        }

    def calculate_draft_priority(self, overall_df: pd.DataFrame) -> pd.Series:
        """Calculate Draft Priority scores for all players."""
        draft_priorities = []

        # Precompute normalization values
        max_vorp = overall_df["VORP"].max()
        min_vorp = overall_df["VORP"].min()
        max_points = overall_df["Points"].max()

        # Precompute position rankings for efficient lookup
        position_rankings = self._calculate_position_rankings(overall_df)

        for i in range(len(overall_df)):
            player = overall_df.iloc[i]
            priority = self._calculate_single_player_priority(
                player, max_vorp, min_vorp, max_points, position_rankings
            )
            draft_priorities.append(round(priority, 1))

        return pd.Series(draft_priorities)

    def _calculate_position_rankings(self, overall_df: pd.DataFrame) -> dict:
        """Calculate position-specific rankings for all players."""
        position_rankings = {}

        for pos in overall_df["position"].unique():
            if pd.isna(pos) or pos == "":
                continue

            pos_players = (
                overall_df[overall_df["position"] == pos]
                .sort_values(["Points", "VORP"], ascending=[False, False])
                .reset_index(drop=True)
            )

            # Create mapping from player name to position rank
            pos_ranks = {}
            for idx, (_, player) in enumerate(pos_players.iterrows(), 1):
                pos_ranks[player["player"]] = idx

            position_rankings[pos] = pos_ranks

        return position_rankings

    def _calculate_single_player_priority(
        self,
        player,
        max_vorp: float,
        min_vorp: float,
        max_points: float,
        position_rankings: dict,
    ) -> float:
        """Calculate Draft Priority for a single player."""
        pos = player["position"]
        vorp = player["VORP"]
        adp_diff = player["ADP_Diff"]
        adp = pd.to_numeric(player["adp"], errors="coerce")
        if pd.isna(adp):
            adp = 999
        rank = player["Rank"]
        player_name = str(player["player"])

        # Get position rank
        pos_rank = position_rankings.get(pos, {}).get(player_name, 999)

        # Base talent score (0-100)
        talent_score = self._calculate_talent_score(
            vorp, player["Points"], max_vorp, min_vorp, max_points
        )

        # Statistical-based round bonuses (no hardcoded names)
        round_bonus = self._calculate_statistical_round_bonus(pos, vorp, rank, pos_rank)

        # ADP Reality Check - temper statistical bonuses with market reality
        adp_adjustment = self._calculate_adp_reality_check(pos, adp, round_bonus)

        # Position-based adjustments
        pos_multiplier = self._get_position_multiplier(pos)

        # Additional bonuses and penalties
        elite_bonus = self._calculate_elite_bonus(rank, vorp)
        value_bonus = self._calculate_value_bonus(adp_diff, vorp)
        reach_penalty = self._calculate_reach_penalty(adp, adp_diff)

        # Penalize negative VORP players unless they have round bonus
        if vorp < 0 and round_bonus == 0:
            talent_score = max(0, talent_score * 0.3)

        # Calculate final Draft Priority Score
        return (
            talent_score * pos_multiplier
            + round_bonus
            + adp_adjustment
            + elite_bonus
            + value_bonus
            + reach_penalty
        )

    def _calculate_talent_score(
        self,
        vorp: float,
        points: float,
        max_vorp: float,
        min_vorp: float,
        max_points: float,
    ) -> float:
        """Calculate base talent score from VORP and points."""
        # Normalize VORP to 0-60 scale
        if max_vorp != min_vorp:
            vorp_score = ((vorp - min_vorp) / (max_vorp - min_vorp)) * 60
        else:
            vorp_score = 30

        # Add points component (0-40 scale) for raw production
        points_score = (points / max_points) * 40 if max_points > 0 else 0

        return vorp_score + points_score

    def _calculate_statistical_round_bonus(
        self, pos: str, vorp: float, overall_rank: int, pos_rank: int
    ) -> float:
        """Calculate round-based priority bonus using statistical thresholds instead of hardcoded names."""

        # Check each round tier in order (most valuable first)
        for round_tier, config in self.strategy_config["round_bonuses"].items():
            bonus = config["bonus"]
            criteria = config["criteria"].get(pos, {})

            if not criteria:
                continue

            # Check if player meets all criteria for this round tier
            meets_criteria = True

            if "min_vorp" in criteria and vorp < criteria["min_vorp"]:
                meets_criteria = False

            if "max_vorp" in criteria and vorp > criteria["max_vorp"]:
                meets_criteria = False

            if "min_rank" in criteria and overall_rank > criteria["min_rank"]:
                meets_criteria = False

            if "max_rank" in criteria and overall_rank > criteria["max_rank"]:
                meets_criteria = False

            if "min_pos_rank" in criteria and pos_rank > criteria["min_pos_rank"]:
                meets_criteria = False

            if "max_pos_rank" in criteria and pos_rank > criteria["max_pos_rank"]:
                meets_criteria = False

            if meets_criteria:
                return bonus

        return 0

    def _get_position_multiplier(self, pos: str) -> float:
        """Get position-based scoring multiplier."""
        if pos == "QB" and self.is_superflex:
            return 1.05  # Very reduced superflex boost
        elif pos == "TE":
            return 1.1  # Small TE scarcity boost
        elif pos in ["DL", "LB", "DB"]:
            return 0.8  # IDP gets lower base multiplier
        return 1.0

    def _calculate_elite_bonus(self, rank: int, vorp: float) -> float:
        """Calculate elite player bonus."""
        if rank <= 5 and vorp > 50:
            return 20
        elif rank <= 12 and vorp > 30:
            return 15
        elif rank <= 24 and vorp > 15:
            return 10
        return 0

    def _calculate_value_bonus(self, adp_diff: float, vorp: float) -> float:
        """Calculate value pick bonus."""
        if adp_diff > 20 and vorp > 5:
            return 8
        elif adp_diff > 10 and vorp > 0:
            return 4
        return 0

    def _calculate_reach_penalty(self, adp: float, adp_diff: float) -> float:
        """Calculate reach penalty for high ADP players."""
        if adp < 50 and adp_diff < -15:
            return -15
        elif adp_diff < -10:
            return -8
        return 0

    def _calculate_adp_reality_check(
        self, pos: str, adp: float, round_bonus: float
    ) -> float:
        """
        ADP Reality Check - Balance statistical performance with market expectations.

        The market (ADP) isn't perfect, but it provides useful signal about true player value.
        This method tempers pure statistical rankings with community consensus.
        """
        if adp >= 999:  # No ADP data available
            return 0

        # Position-specific ADP thresholds for round bonuses
        early_pick_thresholds = {
            "QB": 8,  # Only top 8 QBs should get major bonuses (QB1-QB8)
            "RB": 12,  # RBs taken in top 12 are consensus RB1s
            "WR": 18,  # WRs taken in top 18 are consensus WR1s
            "TE": 30,  # TEs taken in top 30 are startable
            "DL": 60,  # IDP taken in top 60 are consensus starters
            "LB": 60,
            "DB": 80,
        }

        mid_pick_thresholds = {
            "QB": 24,  # QBs after pick 24 should have limited bonuses
            "RB": 24,  # RBs after pick 24 are RB2s/flex
            "WR": 36,  # WRs after pick 36 are WR2s/flex
            "TE": 60,  # TEs after pick 60 are backup/streaming
            "DL": 120,  # IDP after pick 120 are late-round fills
            "LB": 120,
            "DB": 150,
        }

        late_pick_thresholds = {
            "QB": 60,  # QBs after pick 60 are bench/streaming
            "RB": 48,  # RBs after pick 48 are handcuffs/dart throws
            "WR": 72,  # WRs after pick 72 are WR3s/depth
            "TE": 120,  # TEs after pick 120 are waiver wire
            "DL": 180,
            "LB": 180,
            "DB": 200,
        }

        early_threshold = early_pick_thresholds.get(pos, 24)
        mid_threshold = mid_pick_thresholds.get(pos, 48)
        late_threshold = late_pick_thresholds.get(pos, 96)

        # ADP-based adjustments to round bonuses - much more aggressive
        if adp <= early_threshold:
            # Market consensus early pick - validate high bonuses
            if round_bonus >= 35:
                return 0  # Market agrees - no adjustment needed
            elif round_bonus >= 20:
                return 8  # Market suggests higher value
            else:
                return 15  # Market strongly disagrees - boost significantly

        elif adp <= mid_threshold:
            # Market consensus mid-round pick - moderate bonuses appropriate
            if round_bonus >= 35:
                return -25  # Strong penalty for high bonus + mid ADP
            elif round_bonus >= 20:
                return -10  # Moderate penalty for medium bonus + mid ADP
            else:
                return 5  # Slight boost for solid mid-round option

        elif adp <= late_threshold:
            # Late ADP - high bonuses likely wrong
            if round_bonus >= 35:
                return -40  # Major penalty for late ADP + high bonus
            elif round_bonus >= 20:
                return -25  # Strong penalty for late ADP + medium bonus
            else:
                return 0  # No adjustment for appropriately low bonus

        else:
            # Very late/undrafted ADP - severely penalize high bonuses
            if round_bonus >= 35:
                return -50  # Massive penalty for undrafted + high bonus
            elif round_bonus >= 20:
                return -35  # Large penalty for undrafted + medium bonus
            elif round_bonus >= 10:
                return -15  # Moderate penalty for undrafted + low bonus
            else:
                return 0  # No penalty for no bonus

        return 0


class VORPCalculator:
    """Handles VORP (Value Over Replacement Player) calculations."""

    @staticmethod
    def calculate_replacement_rank(pos: str, league_config: dict) -> int:
        """Calculate replacement player rank for a position."""
        teams = league_config["num_teams"]
        bench = league_config["bench_factor"].get(pos, 0.5)
        if pos == "QB" and league_config["superflex"]:
            starters = league_config["superflex_qb_per_team"]
        else:
            starters = league_config["starters"].get(pos, 1)
        return max(1, int(round(teams * (starters + bench))))

    @staticmethod
    def calculate_vorp(pool_df: pd.DataFrame, league_config: dict) -> pd.DataFrame:
        """Calculate VORP for all players in the pool."""
        result_df = pool_df.copy()
        result_df["VORP"] = 0.0

        for pos in result_df["position"].dropna().unique():
            pos_players = (
                result_df[result_df["position"] == pos]
                .sort_values("Points", ascending=False)
                .reset_index(drop=True)
            )

            if pos_players.empty:
                continue

            rep_rank = VORPCalculator.calculate_replacement_rank(pos, league_config)
            rep_points = pos_players.iloc[min(rep_rank - 1, len(pos_players) - 1)][
                "Points"
            ]

            # Map VORP back to original dataframe
            result_df.loc[result_df["position"] == pos, "VORP"] = result_df.loc[
                result_df["position"] == pos, "Points"
            ] - float(rep_points)

        return result_df


class TierAssigner:
    """Handles tier assignment based on point gaps."""

    @staticmethod
    def assign_tiers(sorted_df: pd.DataFrame, gap: float) -> List[int]:
        """Assign tiers based on point drops between consecutive players."""
        tiers = []
        current_tier = 1

        for i in range(len(sorted_df)):
            if i == 0:
                tiers.append(current_tier)
                continue

            point_drop = sorted_df.iloc[i - 1]["Points"] - sorted_df.iloc[i]["Points"]
            if point_drop >= gap:
                current_tier += 1

            tiers.append(current_tier)

        return tiers


# Legacy function for backward compatibility
def replacement_rank(pos: str, cfg: dict) -> int:
    return VORPCalculator.calculate_replacement_rank(pos, cfg)


def assign_tiers(sorted_df: pd.DataFrame, gap: float) -> List[int]:
    return TierAssigner.assign_tiers(sorted_df, gap)


class ExcelFormatter:
    """Handles Excel formatting and styling operations."""

    @staticmethod
    def apply_strikethrough_formatting(
        excel_file_path, worksheet_name, drafted_column="I"
    ):
        """Apply conditional strikethrough formatting to a worksheet when 'X' is entered in the drafted column"""
        try:
            # Load the workbook
            wb = load_workbook(excel_file_path)
            ws = wb[worksheet_name]

            # Get the column number for the drafted column
            drafted_col_num = ord(drafted_column.upper()) - 64  # A=1, B=2, etc.

            # Create the conditional formatting rule
            strikethrough_rule = FormulaRule(
                formula=[f'${drafted_column}2="X"'],
                font=Font(strikethrough=True, color="808080")
            )

            # Apply to columns A through the column before drafted (A-H if drafted is I)
            last_data_col = chr(64 + drafted_col_num - 1)  # H if drafted is I

            # Get the last row with data
            last_row = 2  # Start with row 2 (header is row 1)
            for row in range(2, ws.max_row + 1):
                if ws[f'A{row}'].value is not None:
                    last_row = row

            # Apply conditional formatting to each column A through last_data_col
            for col in range(ord('A'), ord(last_data_col) + 1):
                col_letter = chr(col)
                range_str = f"{col_letter}2:{col_letter}{last_row}"
                ws.conditional_formatting.add(range_str, strikethrough_rule)

            # Save the workbook
            wb.save(excel_file_path)
            print(
                f"✅ Automatic strikethrough formatting applied to {worksheet_name} sheet!"
            )

        except Exception as e:
            print(f"⚠️  Could not apply strikethrough formatting: {e}")

    @staticmethod
    def auto_color_worksheet(ws, tier_col=8):
        """Apply tier-based coloring to worksheet rows."""
        fill_map = {
            1: PatternFill(
                start_color="006400", end_color="006400", fill_type="solid"
            ),  # dark green
            2: PatternFill(
                start_color="90EE90", end_color="90EE90", fill_type="solid"
            ),  # light green
            3: PatternFill(
                start_color="FFF59D", end_color="FFF59D", fill_type="solid"
            ),  # soft yellow
            4: PatternFill(
                start_color="FFE082", end_color="FFE082", fill_type="solid"
            ),  # amber
            5: PatternFill(
                start_color="FFCC80", end_color="FFCC80", fill_type="solid"
            ),  # orange
            6: PatternFill(
                start_color="BDBDBD", end_color="BDBDBD", fill_type="solid"
            ),  # gray
        }

        max_row = ws.max_row
        max_col = ws.max_column

        for r in range(2, max_row + 1):
            try:
                tier_value = int(ws.cell(row=r, column=tier_col).value)
            except Exception:
                continue

            fill = fill_map.get(tier_value)
            if fill:
                for c in range(1, max_col + 1):
                    ws.cell(row=r, column=c).fill = fill

    @staticmethod
    def apply_header_formatting(ws):
        """Apply formatting to header row."""
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def set_column_widths(ws, widths: List[int]):
        """Set column widths for worksheet."""
        for idx, width in enumerate(widths, start=1):
            if idx <= len(widths):
                col_letter = ws.cell(row=1, column=idx).column_letter
                ws.column_dimensions[col_letter].width = width

    @staticmethod
    def apply_value_highlighting(ws, value_col: int, max_row: int):
        """Apply highlighting for value picks and reaches."""
        for row in range(2, max_row + 1):
            value_cell = ws.cell(row=row, column=value_col)

            if value_cell.value == "STEAL":
                color = PatternFill(
                    start_color="90EE90", end_color="90EE90", fill_type="solid"
                )
            elif value_cell.value == "VALUE":
                color = PatternFill(
                    start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
                )
            elif value_cell.value == "REACH":
                color = PatternFill(
                    start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
                )
            else:
                continue

            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = color


# Legacy function for backward compatibility
def auto_color_worksheet(ws, tier_col=8):
    return ExcelFormatter.auto_color_worksheet(ws, tier_col)


# -------------------------------
# 4) LOAD DATA
# -------------------------------


def load_csv_flex(path: str) -> pd.DataFrame:
    # Resolve relative paths relative to the script location
    if not os.path.isabs(path):
        path = os.path.join(os.path.dirname(__file__), path)
    if not os.path.exists(path):
        raise FileNotFoundError(f"Missing file: {path}")
    return pd.read_csv(path)


def load_csv_absolute(path: str) -> pd.DataFrame:
    """Load CSV without path resolution - assumes path is already absolute or correctly resolved."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Missing file: {path}")
    return pd.read_csv(path)


def load_sleeper_adp(cfg: dict) -> pd.DataFrame:
    """
    Primary: load from CSV export (most stable).
    Optional: if CONFIG['sleeper_api']['use_api'] is True, try fetching JSON from Sleeper.
    """
    p = cfg["paths"]["adp_csv"]
    # Resolve relative paths relative to the script location
    if not os.path.isabs(p):
        p = os.path.join(os.path.dirname(__file__), p)
    if os.path.exists(p):
        raw = pd.read_csv(p)
    elif cfg["sleeper_api"].get("use_api", False) and requests is not None:
        url = cfg["sleeper_api"]["source_url"].format(
            season=cfg["sleeper_api"]["season"],
            stype=cfg["sleeper_api"]["season_type"],
            scoring=cfg["sleeper_api"]["scoring"],
        )
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        raw = pd.json_normalize(data)
    else:
        raise FileNotFoundError(
            f"No ADP CSV found at {p}. Provide a CSV export or enable API in CONFIG."
        )
    # Standardize
    adp_std = standardize(raw, ADP_MAP)
    # Keep key columns
    return adp_std[["player", "position", "team", "adp"]]


def load_consensus_projections(cfg: dict) -> pd.DataFrame:
    """
    Load and merge projections from multiple sources using weighted consensus.

    Args:
        cfg: Configuration dictionary containing offense_sources list

    Returns:
        Standardized consolidated projections DataFrame
    """
    offense_sources = cfg["paths"]["offense_sources"]
    consensus_config = cfg.get("consensus", {})

    # Load all source files
    source_dfs = []
    all_players = set()  # Track all unique players across sources

    print(f"📊 Loading {len(offense_sources)} projection sources...")

    for source in offense_sources:
        name = source["name"]
        path = source["path"]
        weight = source.get("weight", 1.0)

        try:
            # Load and standardize the raw data
            raw_df = load_csv_absolute(path)
            std_df = standardize_offense_position_aware(raw_df)

            # Add source identifier for tracking
            std_df["_source"] = name
            std_df["_weight"] = weight

            source_dfs.append(std_df)
            all_players.update(std_df["player"].tolist())

            print(f"  ✅ {name}: {len(std_df)} players (weight: {weight})")

        except Exception as e:
            print(f"  ❌ {name}: Failed to load - {e}")
            continue

    if not source_dfs:
        raise ValueError("No projection sources could be loaded successfully")

    # Get all possible numeric columns from all sources
    all_numeric_cols = set()
    exclude_cols = {"player", "position", "team", "age", "_source", "_weight"}

    for df in source_dfs:
        numeric_cols = [
            col
            for col in df.columns
            if col not in exclude_cols and pd.api.types.is_numeric_dtype(df[col])
        ]
        all_numeric_cols.update(numeric_cols)

    all_numeric_cols = list(all_numeric_cols)

    print(
        f"📈 Merging projections for {len(all_numeric_cols)} statistical categories..."
    )
    print(f"   Found {len(all_players)} unique players across all sources")

    # Create a master list of all players with their positions and teams
    player_info = {}
    for df in source_dfs:
        for _, row in df.iterrows():
            player = row["player"]
            if player not in player_info:
                player_info[player] = {
                    "position": row.get("position", ""),
                    "team": row.get("team", ""),
                    "age": row.get("age", None),
                }

    # Create consensus projections for all players
    consensus_rows = []

    for player, info in player_info.items():
        # Start with player identification
        consensus_row = {
            "player": player,
            "position": info["position"],
            "team": info["team"],
            "age": info["age"],
        }

        # Find all data for this player across sources
        player_data = []
        for df in source_dfs:
            player_rows = df[df["player"] == player]
            if not player_rows.empty:
                player_data.append(player_rows.iloc[0])

        # If we have data for this player from at least one source
        if len(player_data) >= consensus_config.get("min_sources", 1):
            # Calculate consensus values for each numeric column
            for col in all_numeric_cols:
                values = []
                weights = []

                for row in player_data:
                    if col in row and not pd.isna(row[col]):
                        values.append(row[col])
                        weights.append(row["_weight"])

                if values:
                    # Calculate weighted average
                    if len(values) == 1:
                        consensus_row[col] = values[0]
                    else:
                        # Handle outlier detection if configured
                        outlier_threshold = consensus_config.get(
                            "outlier_threshold", None
                        )
                        if outlier_threshold and len(values) > 2:
                            mean_val = np.mean(values)
                            std_val = np.std(values)
                            if std_val > 0:  # Avoid division by zero
                                outlier_mask = (
                                    np.abs(np.array(values) - mean_val)
                                    <= outlier_threshold * std_val
                                )
                                values = np.array(values)[outlier_mask].tolist()
                                weights = np.array(weights)[outlier_mask].tolist()

                        if values and np.sum(weights) > 0:
                            consensus_row[col] = np.average(values, weights=weights)
                        else:
                            consensus_row[col] = 0.0
                else:
                    # No data for this column, use 0
                    consensus_row[col] = 0.0

            consensus_rows.append(consensus_row)

    # Create final consensus DataFrame
    consensus_df = pd.DataFrame(consensus_rows)

    # Ensure all expected columns exist with default values if missing
    expected_numeric_cols = [
        "pass_yards",
        "pass_td",
        "pass_int",
        "pass_2pt",
        "rush_yards",
        "rush_td",
        "rush_2pt",
        "receptions",
        "rec_yards",
        "rec_td",
        "rec_2pt",
    ]

    for col in expected_numeric_cols:
        if col not in consensus_df.columns:
            consensus_df[col] = 0.0

    print(f"✅ Created consensus projections for {len(consensus_df)} players")
    print(f"   Method: {consensus_config.get('method', 'weighted_average')}")
    print(f"   Sources used: {[s['name'] for s in offense_sources]}")

    return consensus_df


# -------------------------------
# 5) MAIN PIPELINE
# -------------------------------


def main(
    config_file: str = "config.json",
    config_name: str = "default",
    strategy_file: str = None,
    strategy_name: str = None,
):
    """
    Main function to generate fantasy cheat sheet.

    Args:
        config_file: Path to JSON configuration file
        config_name: Name of configuration to use
        strategy_file: Optional path to draft strategy configuration
        strategy_name: Optional name of strategy configuration to use
    """
    cfg = load_config(config_file, config_name)

    print(f"📋 Using configuration: '{config_name}' from {config_file}")

    # Load optional draft strategy config
    draft_strategy_config = None
    if strategy_file and os.path.exists(strategy_file):
        try:
            with open(strategy_file, "r") as f:
                strategy_configs = json.load(f)
            if strategy_name and strategy_name in strategy_configs:
                draft_strategy_config = strategy_configs[strategy_name]
                print(
                    f"📈 Using draft strategy: '{strategy_name}' from {strategy_file}"
                )
            else:
                print(
                    f"⚠️ Strategy '{strategy_name}' not found in {strategy_file}, using default statistical thresholds"
                )
        except Exception as e:
            print(
                f"⚠️ Could not load draft strategy file: {e}, using default statistical thresholds"
            )

    # Load ADP (Sleeper)
    adp = load_sleeper_adp(cfg)

    # Load offense projections (always try multi-source first)
    # Check if config already has multi-source sources defined
    if "offense_sources" in cfg["paths"]:
        # Use existing multi-source configuration
        off = load_consensus_projections(cfg)
        print(
            f"✅ Using consensus projections from {len(cfg['paths']['offense_sources'])} sources"
        )
        off_raw = off  # For consistency in later checks
    else:
        # Define default multi-source configuration
        default_offense_sources = [
            {
                "name": "2025_Projections",
                "path": os.path.join(
                    os.path.dirname(__file__),
                    "../../data/projections_2025_offense_only.csv",
                ),
                "weight": 0.50,
            },
            {
                "name": "WR_Rankings",
                "path": os.path.join(
                    os.path.dirname(__file__), "../../data/fantasy2025rankingsexcel.csv"
                ),
                "weight": 0.25,
            },
        ]

        # Add primary source if available in config
        if "offense_csv" in cfg["paths"]:
            primary_path = cfg["paths"]["offense_csv"]
            # Resolve relative paths relative to the script location
            if not os.path.isabs(primary_path):
                primary_path = os.path.join(os.path.dirname(__file__), primary_path)
            default_offense_sources.append(
                {"name": "Primary", "path": primary_path, "weight": 0.25}
            )

        # Check if multi-source files are available
        multi_source_available = True
        for source in default_offense_sources:
            if not os.path.exists(source["path"]):
                multi_source_available = False
                break  # Use multi-source if available, otherwise fall back to single source

        if multi_source_available:
            # Temporarily add offense_sources to config for consensus loading
            original_paths = cfg["paths"].copy()
            cfg["paths"]["offense_sources"] = default_offense_sources

            # Multi-source consensus projections
            off = load_consensus_projections(cfg)
            print(
                f"✅ Using consensus projections from {len(default_offense_sources)} sources"
            )

            # Restore original paths
            cfg["paths"] = original_paths

            # For consensus, we already have standardized data
            off_raw = off  # Use the same data for column checking
        elif "offense_csv" in cfg["paths"]:
            # Single source projections (fallback)
            off_raw = load_csv_flex(cfg["paths"]["offense_csv"])
            off = standardize_offense_position_aware(off_raw)
            print("✅ Using single-source projections (multi-source files not found)")
        else:
            # No projection sources available
            raise FileNotFoundError(
                "No offense projection sources found. Configure either offense_csv or multi-source files."
            )

    # Clean player names (remove quotes if present)
    off["player"] = off["player"].astype(str).str.strip('"')

    # Extract base position from position strings like 'WR1', 'RB2'
    off["position"] = off["position"].apply(extract_position)

    # Check if we have actual projection columns or just rankings
    has_proj_stats = any(
        any(col in off_raw.columns for col in OFFENSE_MAP[stat_col])
        for stat_col in ["pass_yards", "rush_yards", "receptions"]
    )

    if has_proj_stats:
        # Use traditional projection-based scoring
        off["Points"] = apply_offense_scoring(off, cfg["scoring"]["offense"])
        print("✅ Using projection-based scoring for offense")
    else:
        # Use ranking-based scoring
        print("⚠️ No projection columns found, using ranking-based scoring for offense")
        off["Points"] = 0.0
        for pos in off["position"].unique():
            if pos and pos != "":
                pos_mask = off["position"] == pos
                pos_data = off[pos_mask].copy()
                off.loc[pos_mask, "Points"] = assign_points_from_rankings(pos_data, pos)

    off["position"] = off["position"].replace({"FB": "RB"})  # normalize if needed

    # Optional IDP
    idp_path = cfg["paths"]["idp_csv"]
    if not os.path.isabs(idp_path):
        idp_path = os.path.join(os.path.dirname(__file__), idp_path)
    use_idp = cfg["league"]["use_idp"] and os.path.exists(idp_path)
    if use_idp:
        idp_raw = load_csv_flex(cfg["paths"]["idp_csv"])
        idp = standardize(idp_raw, IDP_MAP)

        # Clean player names
        idp["player"] = idp["player"].astype(str).str.strip('"')

        # Extract base position
        idp["position"] = idp["position"].apply(extract_position)

        # Check if we have actual projection columns or just rankings
        has_idp_proj_stats = any(
            any(col in idp_raw.columns for col in IDP_MAP[stat_col])
            for stat_col in ["tackle_solo", "tackles_total", "sack", "int"]
        )

        if has_idp_proj_stats:
            idp["Points"] = apply_idp_scoring(idp, cfg["scoring"]["idp"])
            print("✅ Using projection-based scoring for IDP")
        else:
            print("⚠️ No projection columns found, using ranking-based scoring for IDP")
            idp["Points"] = 0.0
            for pos in idp["position"].unique():
                if pos and pos != "":
                    pos_mask = idp["position"] == pos
                    pos_data = idp[pos_mask].copy()
                    idp.loc[pos_mask, "Points"] = assign_points_from_rankings(
                        pos_data, pos
                    )

        pool = pd.concat([off, idp], ignore_index=True, sort=False)
    else:
        pool = off.copy()

    # Name standardization for common mismatches before ADP merge
    def normalize_player_name(name):
        """Normalize player names by removing common suffixes and standardizing format."""
        if not isinstance(name, str):
            return name

        # Convert to title case for consistency
        name = name.strip().title()

        # Remove common suffixes (Jr., Sr., II, III, etc.)
        suffixes = [
            " Jr.",
            " Sr.",
            " II",
            " III",
            " IV",
            " V",
            " Jr",
            " Sr",
            " Ii",
            " Iii",
            " Iv",
            " V",
            " Jr.",
            " Sr.",
            " Ii.",
            " Iii.",
            " Iv.",
            " V.",
        ]

        for suffix in suffixes:
            if name.endswith(suffix):
                name = name[: -len(suffix)].strip()
                break

        # Handle other common name variations
        import re

        # Remove apostrophes (Ja'Marr -> JaMarr, De'Von -> DeVon)
        name = name.replace("'", "")

        # Remove periods (A.J. -> AJ, C.J. -> CJ)
        name = name.replace(".", "")

        # Standardize hyphens (Amon-Ra -> AmonRa, Jaxon Smith-Njigba -> Jaxon SmithNjigba)
        name = name.replace("-", "")

        # Remove extra spaces and normalize spacing
        name = re.sub(r"\s+", " ", name).strip()

        return name

    # Apply name normalization to both datasets
    adp_copy = adp.copy()
    pool_copy = pool.copy()

    adp_copy["normalized_name"] = adp_copy["player"].apply(normalize_player_name)
    pool_copy["normalized_name"] = pool_copy["player"].apply(normalize_player_name)

    # Merge using normalized names
    pool_copy = pool_copy.merge(
        adp_copy[["normalized_name", "adp"]], on=["normalized_name"], how="left"
    )

    # Update the original pool with the merged ADP data
    pool["adp"] = pool_copy["adp"]

    # Handle duplicate player names (e.g., Josh Allen QB vs DL)
    # For players with same name, keep ADP only for the one with highest projected points
    duplicate_names = pool[pool.duplicated(subset=["player"], keep=False)][
        "player"
    ].unique()

    for dup_name in duplicate_names:
        dup_players = pool[pool["player"] == dup_name].copy()
        if len(dup_players) > 1:
            # Find the player with highest points (most likely the "real" player)
            max_points_idx = dup_players["Points"].idxmax()
            # Set ADP to NaN for all except the highest-points player
            pool.loc[
                (pool["player"] == dup_name) & (pool.index != max_points_idx), "adp"
            ] = pd.NA

    # Fill missing ADP with high numbers (late picks)
    pool["adp"] = pool["adp"].fillna(999)

    # VORP calculation using the new class
    pool = VORPCalculator.calculate_vorp(pool, cfg["league"])

    # Build Overall sheet with enhanced metrics
    overall = pool.sort_values("Points", ascending=False).reset_index(drop=True)
    overall["Tier"] = 0
    overall["Rank"] = range(1, len(overall) + 1)  # Overall ranking
    overall["Draft_Value"] = ""  # For draft value indicators
    overall["ADP_Diff"] = overall["Rank"] - overall["adp"]  # Positive = value pick

    # Calculate Draft Priority Score using the dedicated calculator
    priority_calculator = DraftPriorityCalculator(cfg["league"], draft_strategy_config)
    overall["Draft_Priority"] = priority_calculator.calculate_draft_priority(overall)

    # Sort by Draft Priority (highest first) for the final overall ranking
    overall = overall.sort_values("Draft_Priority", ascending=False).reset_index(
        drop=True
    )
    overall["Draft_Rank"] = range(1, len(overall) + 1)  # New draft-optimized ranking

    # assign tiers for overall using FLEX gap
    gaps = cfg["tiers"]["tier_gap_points"]["FLEX"]
    overall["Tier"] = TierAssigner.assign_tiers(overall, gaps)

    # Add value indicators
    for i in range(len(overall)):
        adp_diff = overall.iloc[i]["ADP_Diff"]
        if adp_diff >= 20:
            overall.iloc[i, overall.columns.get_loc("Draft_Value")] = "STEAL"
        elif adp_diff >= 10:
            overall.iloc[i, overall.columns.get_loc("Draft_Value")] = "VALUE"
        elif adp_diff <= -20:
            overall.iloc[i, overall.columns.get_loc("Draft_Value")] = "REACH"
        elif adp_diff <= -10:
            overall.iloc[i, overall.columns.get_loc("Draft_Value")] = "EARLY"

    # Build position sheets with their tier gaps and enhanced metrics
    by_pos = {}
    for pos in ["QB", "RB", "WR", "TE", "DL", "LB", "DB"]:
        sub = (
            pool[pool["position"] == pos]
            .sort_values("Points", ascending=False)
            .reset_index(drop=True)
        )
        if sub.empty:
            continue

        # Add positional ranking and metrics
        sub["Pos_Rank"] = range(1, len(sub) + 1)
        sub["Next_Drop"] = 0.0  # Points drop to next player

        # Calculate point drops to next player
        for i in range(len(sub) - 1):
            sub.iloc[i, sub.columns.get_loc("Next_Drop")] = (
                sub.iloc[i]["Points"] - sub.iloc[i + 1]["Points"]
            )

        # Calculate tiers using the new TierAssigner
        pgap = cfg["tiers"]["tier_gap_points"].get(pos, 10.0)
        sub["Tier"] = TierAssigner.assign_tiers(sub, pgap)

        # Add positional scarcity indicators
        sub["Scarcity"] = ""
        for i in range(len(sub)):
            if sub.iloc[i]["Pos_Rank"] <= 12:  # Top 12 in position
                if sub.iloc[i]["Next_Drop"] >= pgap * 1.5:
                    sub.iloc[i, sub.columns.get_loc("Scarcity")] = "CLIFF"
                elif sub.iloc[i]["Next_Drop"] >= pgap:
                    sub.iloc[i, sub.columns.get_loc("Scarcity")] = "DROP"

        by_pos[pos] = sub

    # Export Excel with enhanced columns
    out = cfg["paths"]["output_xlsx"]
    # Resolve relative paths relative to the script location
    if not os.path.isabs(out):
        out = os.path.join(os.path.dirname(__file__), out)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # Overall sheet with enhanced columns optimized for drafting
        ov = overall.rename(
            columns={
                "player": "Player",
                "team": "Team",
                "position": "Pos",
                "age": "Age",
            }
        )
        # Reorder columns for optimal draft use - Draft_Priority first for easy sorting
        ov = ov[
            [
                "Draft_Priority",
                "Draft_Rank",
                "Player",
                "Pos",
                "Team",
                "Points",
                "VORP",
                "Tier",
                "adp",
                "ADP_Diff",
                "Draft_Value",
                "Rank",
            ]
        ]
        ov.to_excel(writer, sheet_name="Overall", index=False)

        # Position sheets with enhanced columns
        for pos, dfp in by_pos.items():
            dfp2 = dfp.rename(
                columns={
                    "player": "Player",
                    "team": "Team",
                    "position": "Pos",
                    "age": "Age",
                }
            )
            dfp2 = dfp2[
                [
                    "Pos_Rank",
                    "Player",
                    "Team",
                    "Age",
                    "Points",
                    "VORP",
                    "Tier",
                    "Next_Drop",
                    "Scarcity",
                    "adp",
                ]
            ]
            dfp2.to_excel(writer, sheet_name=pos, index=False)

        # Create Draft Strategy Summary Sheet
        strategy_data = []

        # Position Summary
        strategy_data.append(["POSITION SUMMARY", "", "", ""])
        strategy_data.append(
            ["Position", "Top Tier Size", "Avg VORP Top 12", "Strategy Note"]
        )

        for pos in ["QB", "RB", "WR", "TE"]:
            if pos in by_pos:
                pos_df = by_pos[pos]
                tier_1_count = len(pos_df[pos_df["Tier"] == 1])
                top_12_vorp = (
                    pos_df.head(12)["VORP"].mean()
                    if len(pos_df) >= 12
                    else pos_df["VORP"].mean()
                )

                if pos == "QB" and cfg["league"]["superflex"]:
                    note = f"Superflex: Target early (QBs per team: {cfg['league']['superflex_qb_per_team']})"
                elif tier_1_count <= 3:
                    note = "Scarce top tier - prioritize early"
                elif tier_1_count >= 8:
                    note = "Deep position - can wait"
                else:
                    note = "Moderate scarcity"

                strategy_data.append([pos, tier_1_count, f"{top_12_vorp:.1f}", note])

        strategy_data.append(["", "", "", ""])
        strategy_data.append(["VALUE PICKS (Rank much better than ADP)", "", "", ""])
        value_picks = ov[ov["ADP_Diff"] >= 15].head(10)
        strategy_data.append(["Player", "Position", "Rank vs ADP", "Value"])
        for _, player in value_picks.iterrows():
            strategy_data.append(
                [
                    player["Player"],
                    player["Pos"],
                    f"{player['Rank']} vs {player['adp']:.0f}",
                    player["Draft_Value"],
                ]
            )

        strategy_df = pd.DataFrame(strategy_data)
        strategy_df.to_excel(
            writer, sheet_name="Draft Strategy", index=False, header=False
        )

        # Create optimized Draft Board sheet for live drafting
        draft_board = overall.copy()

        # Rename columns first
        draft_board = draft_board.rename(
            columns={
                "player": "Player",
                "team": "Team",
                "position": "Pos",
            }
        )

        # Create simplified draft board with only essential columns
        essential_columns = [
            "Draft_Priority",
            "Draft_Rank",
            "Player",
            "Pos",
            "Team",
            "Tier",
            "adp",
            "Draft_Value",
        ]
        available_columns = [
            col for col in essential_columns if col in draft_board.columns
        ]

        # Select only available columns
        draft_board = draft_board[available_columns]

        # Rename for cleaner display
        column_renames = {
            "Draft_Priority": "Priority",
            "Draft_Rank": "Rank",
            "adp": "ADP",
            "Draft_Value": "Value",
        }
        draft_board = draft_board.rename(columns=column_renames)

        # Add empty "Drafted" column for marking during live draft
        draft_board["Drafted"] = ""
        draft_board["Round"] = ""
        draft_board["Pick"] = ""

        draft_board.to_excel(writer, sheet_name="Draft Board", index=False)

    # Add colors & formatting with enhanced features
    wb = load_workbook(out)
    for name in wb.sheetnames:
        ws = wb[name]
        # header
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        # Color rows by Tier (adjust column based on sheet)
        if name == "Overall":
            auto_color_worksheet(ws, tier_col=8)  # Tier column for Overall
            # Set column widths for Overall sheet
            widths = [
                12,
                8,
                22,
                6,
                8,
                10,
                10,
                6,
                8,
                10,
                12,
                8,
            ]  # Updated for new column order

            # Highlight value picks and reaches in Draft_Value column (column 11)
            for row in range(2, ws.max_row + 1):
                value_cell = ws.cell(row=row, column=11)  # Draft_Value column
                if value_cell.value == "STEAL":
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="90EE90", end_color="90EE90", fill_type="solid"
                        )
                elif value_cell.value == "VALUE":
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
                        )
                elif value_cell.value == "REACH":
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
                        )

            # Highlight top Draft Priority scores (top 24 picks get gold highlighting)
            for row in range(2, min(26, ws.max_row + 1)):  # Top 24 picks
                draft_rank_cell = ws.cell(row=row, column=2)  # Draft_Rank column
                if draft_rank_cell.value <= 24:
                    priority_cell = ws.cell(row=row, column=1)  # Draft_Priority column
                    priority_cell.fill = PatternFill(
                        start_color="FFD700", end_color="FFD700", fill_type="solid"
                    )
                    priority_cell.font = Font(bold=True)

        elif name == "Draft Board":
            # Special formatting for Draft Board
            # Set column widths for Draft Board
            widths = [10, 6, 22, 6, 8, 6, 8, 12, 12, 8, 6]

            # Color by tiers and value
            for row in range(2, ws.max_row + 1):
                tier_cell = ws.cell(row=row, column=6)  # Tier column
                value_cell = ws.cell(row=row, column=8)  # Value column

                # Tier-based row coloring (lighter)
                tier_value = tier_cell.value if tier_cell.value else 1
                if tier_value == 1:
                    row_color = "F0F8FF"  # Light blue for tier 1
                elif tier_value == 2:
                    row_color = "F5F5F5"  # Light gray for tier 2
                elif tier_value <= 4:
                    row_color = "FFFAFA"  # Very light for tiers 3-4
                else:
                    row_color = "FFFFFF"  # White for lower tiers

                # Apply tier coloring to entire row
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color=row_color, end_color=row_color, fill_type="solid"
                    )

                # Override with value colors if applicable
                if value_cell.value == "STEAL":
                    for col in range(1, 9):  # Don't color draft tracking columns
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="90EE90", end_color="90EE90", fill_type="solid"
                        )
                elif value_cell.value == "VALUE":
                    for col in range(1, 9):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
                        )

            # Apply strikethrough formatting to Draft Board using the reusable function
            # This will be done after the Excel file is saved
            pass

        elif name == "Draft Strategy":
            # Set column widths for Draft Strategy
            widths = [25, 15, 15, 40]

        else:
            # Default position sheet formatting
            auto_color_worksheet(ws, tier_col=7)  # Tier column for position sheets
            # Set column widths for position sheets
            widths = [8, 22, 8, 6, 10, 10, 6, 10, 12, 8]  # Updated for new columns

            # Highlight scarcity indicators
            for row in range(2, ws.max_row + 1):
                scarcity_cell = ws.cell(row=row, column=9)  # Scarcity column
                if scarcity_cell.value == "CLIFF":
                    scarcity_cell.fill = PatternFill(
                        start_color="FF4500", end_color="FF4500", fill_type="solid"
                    )
                    scarcity_cell.font = Font(color="FFFFFF", bold=True)
                elif scarcity_cell.value == "DROP":
                    scarcity_cell.fill = PatternFill(
                        start_color="FFA500", end_color="FFA500", fill_type="solid"
                    )
                    scarcity_cell.font = Font(bold=True)

        # freeze header
        ws.freeze_panes = "A2"
        # set column widths
        for idx, w in enumerate(widths, start=1):
            if idx <= len(widths):
                col_letter = ws.cell(row=1, column=idx).column_letter
                ws.column_dimensions[col_letter].width = w
    wb.save(out)

    print(f"✅ Cheat sheet created: {out}")

    # Apply strikethrough formatting to Draft Board sheet
    # Normalize the path for Excel
    normalized_path = os.path.abspath(out)
    ExcelFormatter.apply_strikethrough_formatting(normalized_path, "Draft Board", "I")

    # Print draft insights using the new printer class
    DraftInsightsPrinter.print_draft_insights(overall, by_pos, use_idp)


class DraftInsightsPrinter:
    """Handles printing draft insights and summary information."""

    @staticmethod
    def print_draft_insights(overall_df: pd.DataFrame, by_pos: dict, use_idp: bool):
        """Print comprehensive draft insights to console."""
        print("\n📊 DRAFT INSIGHTS:")
        print("=" * 50)

        # Show top draft priorities
        DraftInsightsPrinter._print_top_priorities(overall_df)

        # Position scarcity summary
        DraftInsightsPrinter._print_position_scarcity(by_pos)

        # Top value picks
        DraftInsightsPrinter._print_value_picks(overall_df)

        print("=" * 50)
        print(
            "💡 TIP: Sort the 'Overall' sheet by 'Draft_Priority' column (highest first) for optimal draft order!"
        )

        if use_idp:
            print("Includes IDP positions (DL/LB/DB).")
        else:
            print(
                "No IDP included (enable 'use_idp' in config and provide idp_projections.csv)."
            )

        print("🎯 DRAFT BOARD FEATURE: Use the 'Draft Board' sheet for live drafting!")
        print("   - Enter 'X' in the 'Drafted' column to cross out drafted players")
        print("   - See docs/excel_strikethrough_guide.md for auto-strikethrough setup")

        print(
            "💡 Tip: Edit config files to adjust scoring, tier gaps, and team settings."
        )

    @staticmethod
    def _print_top_priorities(overall_df: pd.DataFrame):
        """Print top 10 draft priorities."""
        print("🎯 TOP 10 DRAFT PRIORITIES:")
        top_priorities = overall_df.head(10)

        for i, (_, player) in enumerate(top_priorities.iterrows(), 1):
            print(
                f"  {i:2d}. {player['player']:20s} ({player['position']}) - Priority: {player['Draft_Priority']:5.1f}"
            )

    @staticmethod
    def _print_position_scarcity(by_pos: dict):
        """Print position scarcity summary."""
        print("\n📈 POSITION SCARCITY:")

        for pos in ["QB", "RB", "WR", "TE"]:
            if pos in by_pos:
                pos_df = by_pos[pos]
                tier_1_count = len(pos_df[pos_df["Tier"] == 1])
                print(f"  {pos}: Tier 1 has {tier_1_count} players")

    @staticmethod
    def _print_value_picks(overall_df: pd.DataFrame):
        """Print best value picks."""
        print("\n💎 BEST VALUE PICKS (High Priority + Good ADP):")

        value_priorities = overall_df[
            (overall_df["ADP_Diff"] >= 10) & (overall_df["Draft_Priority"] >= 70)
        ].head(5)

        if not value_priorities.empty:
            for _, player in value_priorities.iterrows():
                print(
                    f"  {player['player']} ({player['position']}) - Priority {player['Draft_Priority']:.1f}, "
                    f"Ranked #{player['Draft_Rank']} vs ADP {player['adp']:.0f}"
                )
        else:
            print("  No major value picks found in current data")


if __name__ == "__main__":
    import sys

    # Parse command-line arguments
    if len(sys.argv) >= 3:
        CONFIG_FILE = sys.argv[1]
        CONFIG_PROFILE = sys.argv[2]
        # Resolve relative config file paths
        if not os.path.isabs(CONFIG_FILE):
            CONFIG_FILE = os.path.join(os.path.dirname(__file__), "../..", CONFIG_FILE)

    # Special command: show all available configurations
    if CONFIG_FILE == "list" and CONFIG_PROFILE == "all":
        list_all_available_configs()
    else:
        # Use configuration (from command-line or defaults at top of file)
        main(CONFIG_FILE, CONFIG_PROFILE)
