import openpyxl

print("OpenPyXL version:", openpyxl.__version__)

# Try different conditional formatting approaches
from openpyxl import load_workbook
from openpyxl.styles import Font

# Load original file
wb = load_workbook("../output/dynasty_superflex_cheatsheet.xlsx")
ws = wb["Draft Board"]

print("Testing different conditional formatting approaches...")

# Method 1: Simple manual approach - create sample rows
print("Creating sample rows with manual strikethrough...")

# Add sample data to show what it should look like
sample_row = 1500
ws.cell(row=sample_row, column=3).value = "SAMPLE: Normal Row"
ws.cell(row=sample_row, column=9).value = ""

sample_row_x = 1501
ws.cell(row=sample_row_x, column=3).value = "SAMPLE: Crossed Out Row"
ws.cell(row=sample_row_x, column=9).value = "X"

# Apply manual strikethrough to sample row
for col in range(1, 9):  # A-H
    cell = ws.cell(row=sample_row_x, column=col)
    cell.font = Font(strikethrough=True, color="808080")

print("Sample rows created")

# Method 2: Try basic conditional formatting
try:
    from openpyxl.conditional_formatting import ConditionalFormatting
    from openpyxl.formatting.rule import Rule

    # Create a simple rule
    rule = Rule(type="expression", formula=['$I2="X"'], font=Font(strikethrough=True))

    # Apply to a small range first
    ws.conditional_formatting.add("A2:H10", rule)
    print("Basic conditional formatting applied to A2:H10")

except Exception as e:
    print(f"Conditional formatting failed: {e}")

# Save with new name
wb.save("../output/dynasty_superflex_cheatsheet_manual_strikethrough.xlsx")
print("File saved as: dynasty_superflex_cheatsheet_manual_strikethrough.xlsx")
