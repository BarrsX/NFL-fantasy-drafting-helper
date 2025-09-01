import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Load the workbook
wb = load_workbook("../output/dynasty_superflex_cheatsheet.xlsx")
ws = wb["Draft Board"]

print("Applying conditional formatting...")

# Method 1: Try using CellIsRule (simpler approach)
try:
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font

    # Create strikethrough font
    strike_font = Font(strikethrough=True, color="808080")

    # Apply conditional formatting using CellIsRule
    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        rule = CellIsRule(operator="equal", formula=['"X"'], font=strike_font)
        ws.conditional_formatting.add(f"{col}2:{col}1500", rule)

    print("✅ Applied CellIsRule conditional formatting")

except ImportError:
    print("❌ CellIsRule not available, trying manual approach...")

    # Method 2: Manual approach - add a test row to demonstrate
    # This will show the user what it should look like
    test_row = 1500  # Use a high row number for testing

    # Add a test entry
    ws.cell(row=test_row, column=9).value = "X"  # Drafted column
    ws.cell(row=test_row, column=3).value = "TEST PLAYER"

    # Manually apply strikethrough to show what it should look like
    for col in range(1, 9):  # Columns A-H
        cell = ws.cell(row=test_row, column=col)
        cell.font = Font(strikethrough=True, color="808080")

    print("✅ Added test row with manual strikethrough formatting")

# Save the workbook with a new name
wb.save("../output/dynasty_superflex_cheatsheet_strikethrough_fixed.xlsx")
print("📁 File saved as: dynasty_superflex_cheatsheet_strikethrough_fixed.xlsx")

# Also create a version with manual instructions
wb2 = load_workbook("../output/dynasty_superflex_cheatsheet.xlsx")
ws2 = wb2["Draft Board"]

# Add clear instructions in the header
ws2.cell(row=1, column=9).value = "Drafted (Type X to cross out row)"

# Add a sample row to show what it should look like
ws2.cell(row=1501, column=3).value = "SAMPLE: Type X in column I"
ws2.cell(row=1501, column=9).value = "X"

# Apply manual strikethrough to sample row
for col in range(1, 9):
    cell = ws2.cell(row=1501, column=col)
    cell.font = Font(strikethrough=True, color="808080")

wb2.save("../output/dynasty_superflex_cheatsheet_demo.xlsx")
print("📁 Demo file saved as: dynasty_superflex_cheatsheet_demo.xlsx")
print("🎯 Check row 1501 to see how strikethrough should look!")
